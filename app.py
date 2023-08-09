import pandas as pd
import geoip2.database
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Carga de archivo csv
datos = pd.read_csv('Prueba de capacidades técnicas - Listado de participantes.csv', header=0, sep=';')

reader = geoip2.database.Reader('GeoLite2-Country.mmdb')

#Funcion obtener pais segun dirección IP
def obtener_pais(ip):
    try:
        response = reader.country(ip)
        country_name = response.country.name
        return country_name
    except Exception as e:
        return "Desconocido"

#Creacion columna Pais y asignacion de datos
datos['Pais'] = datos['GDPR IP'].apply(obtener_pais)

reader.close()

#Crear archivo Excel usando openpyxl
workbook = Workbook()
worksheet = workbook.active

#Agregar encabezados
headers = list(datos.columns)
worksheet.append(headers)

#Agregar datos
for index, row in datos.iterrows():
    worksheet.append(row.tolist())  #Convertir fila a lista y agregar al archivo


#Ajuta el ancho de las columnas según el contenido
for column_index, column in enumerate(worksheet.columns):
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[get_column_letter(column_index + 1)].width = adjusted_width

# Guardar archivo Excel
workbook.save('Listado_de_participantes.xlsx')
